# MIT License
#
# Copyright (c) 2022 Josh Reginaldo (https://github.com/AnotherCreator)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NON-INFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
    THIS FILE WILL HANDLE CREATING THE ACTUAL WORKBOOK AND WORKSHEETS
"""
from src.constants import \
    LOGGER, MONTHS
import src.modules.xlsx_formatting as xlsx_formatting

import os.path
from os.path import exists
import xlsxwriter  # Library link: https://xlsxwriter.readthedocs.io/index.html
from openpyxl import *


"""
    CREATING THE WORKBOOK BASED ON USER YEAR
"""
def init_workbook(input_year):
    # Check if file was already created
    """ This section is for the macro-enabled master sheet style workbook
    logging.info("Attempting to create: " + str(input_year) + " Study Room Log.xlsm")
    if exists("Test " + str(input_year) + " Study Room Log.xlsm"):
        logging.info("Existing file found, file not created")
        logging.info("Attempting to fetch file: "
                     + os.path.basename(str(input_year) + " Study Room Log.xlsm"))

        existing_file = os.path.basename(str(input_year) + " Study Room Log.xlsm")
        WORKBOOK = load_workbook(existing_file)

        logging.info("File successfully fetched")
        logging.info("Leaving function: create_excel_workbook()")
        return WORKBOOK
    # Create Excel file if it does not exist
    else:
        logging.info(str(input_year) + " Study Room Log.xlsm NOT FOUND")
        logging.info("Creating: " + str(input_year) + " Study Room Log.xlsm")
        WORKBOOK = xlsxwriter.Workbook(str(input_year) + " Study Room Log.xlsm")

        WORKBOOK.add_vba_project('vbaProject.bin')  # Add .bin file with preloaded macros | enables .xlsm creation
        master_sheet = WORKBOOK.add_worksheet("Master Worksheet")

        # Url Test Block
        # TODO: Delete this block after automating url writing (148 - 152)
        master_sheet.write("A1", "January")
        master_sheet.write_url("A2", "internal:'Sun Jan 01'!C3", string='Sun Jan 01')
        master_sheet.write_url("A3", "internal:'Mon Jan 02'!C3", string='Mon Jan 02')
        master_sheet.write_url("A4", "internal:'Tue Jan 03'!C3", string='Tue Jan 03')
        master_sheet.set_first_sheet()  # First visible sheet upon opening file
        master_sheet.activate()  # First visible sheet upon opening file
    """

    LOGGER.info("Attempting to create: " + str(input_year) + " Study Room Log.xlsx")
    if exists("Test " + str(input_year) + " Study Room Log.xlsx"):
        LOGGER.info("Existing file found, file not created")
        LOGGER.info("Attempting to fetch file: "
                     + os.path.basename(str(input_year) + " Study Room Log.xlsx"))

        existing_file = os.path.basename(str(input_year) + " Study Room Log.xlsx")
        wb = load_workbook(existing_file)

        LOGGER.info("File successfully fetched")
        LOGGER.info("Leaving function: create_excel_workbook()")
        return wb
    # Create Excel file if it does not exist
    else:
        LOGGER.info(str(input_year) + " Study Room Log.xlsx NOT FOUND")
        LOGGER.info("Creating: " + str(input_year) + " Study Room Log.xlsx")
        wb = xlsxwriter.Workbook(str(input_year) + " Study Room Log.xlsx")
        LOGGER.info("Leaving function: create_excel_workbook()")
        return wb


def create_daily_worksheets(wb, list_of_dates):
    LOGGER.info("Adding '[DayName] [Month] [DayNumber]' sheets")
    # Add 'DayName Month DayNumber' sheets
    # This will iterate through all known dates of the year
    for date in list_of_dates:
        string_date = date.strftime("%a %b %d")
        ws = wb.add_worksheet(string_date)  # Add 'DayName Month DayNumber' sheets
        # Adjust column widths
        ws.set_column(0, 0, 7.5)  # Hourly time width
        ws.set_column(1, 1, 5.5)  # Quarter intervals

        # General worksheet formatting
        """  This section is related to the master sheet style workbook
        # ws.hide()  # Hide worksheet by default -- access by master link
        """
        ws.set_default_row(hide_unused_rows=True)
        xlsx_formatting.create_worksheet_headers(wb, ws)  # Add room columns and formatting
        xlsx_formatting.create_formulas(ws)  # Add formulas

        # Create y-axis time blocks
        if "Sun" in string_date:  # Create time format for Sundays
            if "Jun" in string_date or "Jul" in string_date or "Aug" in string_date:
                xlsx_formatting.create_summer_sun_format(wb, ws)
            else:
                xlsx_formatting.create_sun_format(wb, ws)
        elif "Sat" in string_date:  # Create time format for Saturdays
            xlsx_formatting.create_sat_format(wb, ws)
        else:
            xlsx_formatting.create_week_day_format(wb, ws)  # Create time format for weekdays


def create_month_total_worksheets(wb, list_of_dates):
    LOGGER.info("Adding '[Month] Total' sheets")
    # Add monthly total sheets
    for month in MONTHS:
        ws = wb.add_worksheet(month + " Totals")
        xlsx_formatting.create_month_total_format(wb, ws, list_of_dates, month)


def create_year_total_worksheet(wb, input_year):
    LOGGER.info("Adding '[Year] Total' sheet")
    # Add yearly total sheet at the end
    wb.add_worksheet(str(input_year) + " Totals")
    # TODO: ADD FORMATTING / FORMULAS FOR MONTHLY TOTAL USERS AND YEAR GRAND TOTAL
